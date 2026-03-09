"""
Train Slot Extractor (NER)
Fine-tunes bert-base-multilingual-uncased for BIO token classification.

Dataset: shopping_queries_dataset/slotannotationdataset.xlsx
Tags:    B-PRODUCT1, I-PRODUCT1, B-PRODUCT2, B-CONN, B-COLOR, B-BRAND, 
         B-PRICE_MOD, B-PRICE_MIN, B-PRICE_MAX, B-RATING_MIN, B-RATING_MOD, etc.
Output:  models/slot_extractor/model.pt + tag_map.json + evaluation_results.json

Usage:
  python train_slot_extractor.py
  python train_slot_extractor.py --epochs 10 --batch_size 16 --lr 3e-5
"""

import os
import json
import ast
import torch
import torch.nn as nn
import numpy as np
import openpyxl
from torch.utils.data import DataLoader, Dataset
from transformers import BertModel, BertTokenizerFast, get_linear_schedule_with_warmup
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    classification_report, f1_score, precision_score,
    recall_score, accuracy_score, confusion_matrix
)
from nltk.translate.bleu_score import corpus_bleu, SmoothingFunction
import argparse
import time
from collections import Counter


# ============================================================
#  Model Architecture
# ============================================================

class SlotExtractor(nn.Module):
    """
    BERT + token-level classification head for NER/slot extraction.
    
    Architecture:
      Input text → BERT tokenizer (subword tokens)
      → BERT encoder (768-dim per subword token)
      → Dropout → Linear(768, num_tags)
    
    Each token gets a BIO tag prediction.
    Subword tokens inherit the label of their first piece.
    """
    
    def __init__(self, bert_model_name='bert-base-multilingual-uncased', num_tags=20):
        super().__init__()
        self.bert = BertModel.from_pretrained(bert_model_name)
        self.dropout = nn.Dropout(0.3)
        self.classifier = nn.Linear(768, num_tags)
    
    def forward(self, input_ids, attention_mask):
        outputs = self.bert(input_ids=input_ids, attention_mask=attention_mask)
        sequence_output = outputs.last_hidden_state  # (batch, seq_len, 768)
        sequence_output = self.dropout(sequence_output)
        logits = self.classifier(sequence_output)     # (batch, seq_len, num_tags)
        return logits


# ============================================================
#  Dataset
# ============================================================

class SlotDataset(Dataset):
    """Custom dataset that handles word-to-subword alignment for NER."""
    
    def __init__(self, tokens_list, labels_list, tokenizer, tag2id, max_length=128):
        self.encodings = []
        self.label_ids = []
        
        ignore_id = -100  # PyTorch CrossEntropyLoss ignores this label
        
        for tokens, labels in zip(tokens_list, labels_list):
            # Tokenize word by word to track alignment
            encoding = tokenizer(
                tokens,
                is_split_into_words=True,
                padding='max_length',
                truncation=True,
                max_length=max_length,
                return_tensors='pt'
            )
            
            word_ids = encoding.word_ids(batch_index=0)
            
            # Align labels: first subword gets the label, rest get -100
            aligned_labels = []
            previous_word_idx = None
            for word_idx in word_ids:
                if word_idx is None:
                    aligned_labels.append(ignore_id)  # [CLS], [SEP], [PAD]
                elif word_idx != previous_word_idx:
                    aligned_labels.append(tag2id.get(labels[word_idx], tag2id['O']))
                else:
                    aligned_labels.append(ignore_id)  # subword continuation
                previous_word_idx = word_idx
            
            self.encodings.append({
                'input_ids': encoding['input_ids'].squeeze(0),
                'attention_mask': encoding['attention_mask'].squeeze(0),
            })
            self.label_ids.append(torch.tensor(aligned_labels, dtype=torch.long))
    
    def __len__(self):
        return len(self.encodings)
    
    def __getitem__(self, idx):
        return (
            self.encodings[idx]['input_ids'],
            self.encodings[idx]['attention_mask'],
            self.label_ids[idx]
        )


# ============================================================
#  Data Loading
# ============================================================

def load_slot_dataset(filepath):
    """Load slot annotation dataset from xlsx."""
    print(f"  Loading {filepath}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    tokens_list = []
    labels_list = []
    errors = 0
    
    for r in range(2, ws.max_row + 1):
        raw_tokens = ws.cell(r, 1).value
        raw_labels = ws.cell(r, 2).value
        
        if not raw_tokens or not raw_labels:
            continue
        
        try:
            tokens = ast.literal_eval(str(raw_tokens))
            labels = ast.literal_eval(str(raw_labels))
            
            if len(tokens) != len(labels):
                errors += 1
                continue
            
            tokens_list.append(tokens)
            labels_list.append(labels)
        except Exception:
            errors += 1
            continue
    
    if errors > 0:
        print(f"  ⚠ Skipped {errors} rows with errors")
    
    # Collect all unique tags
    all_tags = Counter()
    for labels in labels_list:
        all_tags.update(labels)
    
    print(f"  Loaded {len(tokens_list)} examples")
    print(f"  Unique tags: {len(all_tags)}")
    print(f"  Tag distribution:")
    for tag, count in sorted(all_tags.items(), key=lambda x: -x[1]):
        print(f"    {tag:25s}: {count}")
    
    return tokens_list, labels_list, sorted(all_tags.keys())


# ============================================================
#  Evaluation
# ============================================================

def collect_predictions(model, val_loader, device, id2tag):
    """Run model on validation set and collect token-level predictions."""
    model.eval()
    all_preds = []
    all_labels = []
    
    with torch.no_grad():
        for input_ids, attention_mask, labels in val_loader:
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            
            logits = model(input_ids, attention_mask)
            preds = torch.argmax(logits, dim=-1).cpu().numpy()
            labels_np = labels.numpy()
            
            # Only collect non-ignored labels (skip [CLS], [SEP], [PAD], subwords)
            for pred_seq, label_seq in zip(preds, labels_np):
                for p, l in zip(pred_seq, label_seq):
                    if l != -100:
                        all_preds.append(p)
                        all_labels.append(l)
    
    pred_tags = [id2tag[p] for p in all_preds]
    true_tags = [id2tag[l] for l in all_labels]
    
    return true_tags, pred_tags


def evaluate_slot_sequences(model, val_loader, device, id2tag):
    """
    Evaluate at SEQUENCE level — reconstruct full entity sequences per query
    and compute entity-level precision/recall/F1.
    """
    model.eval()
    
    # Collect per-sequence predictions for entity-level eval
    all_seq_true = []
    all_seq_pred = []
    
    with torch.no_grad():
        for input_ids, attention_mask, labels in val_loader:
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            
            logits = model(input_ids, attention_mask)
            preds = torch.argmax(logits, dim=-1).cpu().numpy()
            labels_np = labels.numpy()
            
            for pred_seq, label_seq in zip(preds, labels_np):
                seq_true = []
                seq_pred = []
                for p, l in zip(pred_seq, label_seq):
                    if l != -100:
                        seq_true.append(id2tag[l])
                        seq_pred.append(id2tag[p])
                all_seq_true.append(seq_true)
                all_seq_pred.append(seq_pred)
    
    return all_seq_true, all_seq_pred


def extract_entities(tag_sequence, token_sequence=None):
    """
    Extract entity spans from a BIO tag sequence.
    Returns list of (entity_type, start_idx, end_idx) tuples.
    """
    entities = []
    current_entity = None
    start_idx = 0
    
    for i, tag in enumerate(tag_sequence):
        if tag.startswith('B-'):
            if current_entity:
                entities.append((current_entity, start_idx, i))
            current_entity = tag[2:]
            start_idx = i
        elif tag.startswith('I-'):
            entity_type = tag[2:]
            if current_entity != entity_type:
                if current_entity:
                    entities.append((current_entity, start_idx, i))
                current_entity = entity_type
                start_idx = i
        else:  # O tag
            if current_entity:
                entities.append((current_entity, start_idx, i))
                current_entity = None
    
    if current_entity:
        entities.append((current_entity, start_idx, len(tag_sequence)))
    
    return entities


def entity_level_metrics(all_seq_true, all_seq_pred):
    """
    Compute entity-level precision, recall, F1.
    An entity is correct only if BOTH the type and span boundaries match exactly.
    """
    tp = 0
    fp = 0
    fn = 0
    
    entity_tp = Counter()
    entity_fp = Counter()
    entity_fn = Counter()
    
    for true_seq, pred_seq in zip(all_seq_true, all_seq_pred):
        true_entities = set(extract_entities(true_seq))
        pred_entities = set(extract_entities(pred_seq))
        
        matched = true_entities & pred_entities
        tp += len(matched)
        fp += len(pred_entities - true_entities)
        fn += len(true_entities - pred_entities)
        
        for etype, _, _ in matched:
            entity_tp[etype] += 1
        for etype, _, _ in (pred_entities - true_entities):
            entity_fp[etype] += 1
        for etype, _, _ in (true_entities - pred_entities):
            entity_fn[etype] += 1
    
    precision = tp / (tp + fp) if (tp + fp) > 0 else 0.0
    recall = tp / (tp + fn) if (tp + fn) > 0 else 0.0
    f1 = 2 * precision * recall / (precision + recall) if (precision + recall) > 0 else 0.0
    
    # Per-entity-type metrics
    all_entity_types = sorted(set(list(entity_tp.keys()) + list(entity_fp.keys()) + list(entity_fn.keys())))
    per_entity = {}
    for etype in all_entity_types:
        e_tp = entity_tp.get(etype, 0)
        e_fp = entity_fp.get(etype, 0)
        e_fn = entity_fn.get(etype, 0)
        e_prec = e_tp / (e_tp + e_fp) if (e_tp + e_fp) > 0 else 0.0
        e_rec = e_tp / (e_tp + e_fn) if (e_tp + e_fn) > 0 else 0.0
        e_f1 = 2 * e_prec * e_rec / (e_prec + e_rec) if (e_prec + e_rec) > 0 else 0.0
        per_entity[etype] = {
            'precision': round(e_prec, 4),
            'recall': round(e_rec, 4),
            'f1': round(e_f1, 4),
            'support': e_tp + e_fn,
        }
    
    return {
        'precision': round(precision, 4),
        'recall': round(recall, 4),
        'f1': round(f1, 4),
        'total_true_entities': tp + fn,
        'total_pred_entities': tp + fp,
        'correct_entities': tp,
        'per_entity_type': per_entity,
    }


def evaluate_slot_model(model, val_loader, device, id2tag, tag2id, save_dir):
    """
    Comprehensive evaluation of slot extractor.
    
    Metrics computed:
    - Token-Level: Accuracy, Precision, Recall, F1 (per-tag and micro/macro)
    - Entity-Level: Precision, Recall, F1 (strict span matching)
    - Per-tag Confusion Matrix summary
    """
    
    # ---- Token-level metrics ----
    true_tags, pred_tags = collect_predictions(model, val_loader, device, id2tag)
    
    entity_tag_names = sorted(set(tag for tag in id2tag.values() if tag != 'O'))
    all_tag_names = sorted(id2tag.values())
    
    # Token accuracy (including O)
    token_accuracy = accuracy_score(true_tags, pred_tags)
    
    # Micro/Macro for entity tags only (excluding O)
    micro_precision = precision_score(true_tags, pred_tags, labels=entity_tag_names,
                                       average='micro', zero_division=0)
    micro_recall = recall_score(true_tags, pred_tags, labels=entity_tag_names,
                                 average='micro', zero_division=0)
    micro_f1 = f1_score(true_tags, pred_tags, labels=entity_tag_names,
                         average='micro', zero_division=0)
    
    macro_precision = precision_score(true_tags, pred_tags, labels=entity_tag_names,
                                       average='macro', zero_division=0)
    macro_recall = recall_score(true_tags, pred_tags, labels=entity_tag_names,
                                 average='macro', zero_division=0)
    macro_f1 = f1_score(true_tags, pred_tags, labels=entity_tag_names,
                         average='macro', zero_division=0)
    
    weighted_f1 = f1_score(true_tags, pred_tags, labels=entity_tag_names,
                            average='weighted', zero_division=0)
    
    # Per-tag metrics
    per_tag = {}
    for tag in all_tag_names:
        true_binary = [1 if t == tag else 0 for t in true_tags]
        pred_binary = [1 if p == tag else 0 for p in pred_tags]
        support = sum(true_binary)
        if support > 0:
            per_tag[tag] = {
                'precision': round(precision_score(true_binary, pred_binary, zero_division=0), 4),
                'recall': round(recall_score(true_binary, pred_binary, zero_division=0), 4),
                'f1': round(f1_score(true_binary, pred_binary, zero_division=0), 4),
                'accuracy': round(accuracy_score(true_binary, pred_binary), 4),
                'support': support,
            }
    
    # ---- Entity-level metrics (strict span matching) ----
    all_seq_true, all_seq_pred = evaluate_slot_sequences(model, val_loader, device, id2tag)
    entity_metrics = entity_level_metrics(all_seq_true, all_seq_pred)
    
    # ---- BLEU Score ----
    # Compares predicted BIO tag sequences against true tag sequences per query
    smoother = SmoothingFunction().method1
    bleu_references = []
    bleu_hypotheses = []
    for true_seq, pred_seq in zip(all_seq_true, all_seq_pred):
        if len(true_seq) > 0:
            bleu_references.append([true_seq])  # corpus_bleu expects list of references
            bleu_hypotheses.append(pred_seq)
    
    bleu_score = corpus_bleu(bleu_references, bleu_hypotheses, smoothing_function=smoother)
    
    # ---- Build results ----
    results = {
        'token_level': {
            'accuracy': round(token_accuracy, 4),
            'bleu_score': round(bleu_score, 4),
            'micro_precision': round(micro_precision, 4),
            'micro_recall': round(micro_recall, 4),
            'micro_f1': round(micro_f1, 4),
            'macro_precision': round(macro_precision, 4),
            'macro_recall': round(macro_recall, 4),
            'macro_f1': round(macro_f1, 4),
            'weighted_f1': round(weighted_f1, 4),
        },
        'entity_level': entity_metrics,
        'per_tag': per_tag,
        'total_tokens_evaluated': len(true_tags),
        'total_sequences_evaluated': len(all_seq_true),
    }
    
    # ---- Print Report ----
    print("\n" + "=" * 70)
    print("  EVALUATION RESULTS — SLOT EXTRACTOR (NER)")
    print("=" * 70)
    
    print(f"\n  Token-Level Metrics (per-token accuracy):")
    print(f"    Token Accuracy   : {token_accuracy:.4f}")
    print(f"    BLEU Score       : {bleu_score:.4f}")
    print(f"    Micro Precision  : {micro_precision:.4f}")
    print(f"    Micro Recall     : {micro_recall:.4f}")
    print(f"    Micro F1         : {micro_f1:.4f}")
    print(f"    Macro Precision  : {macro_precision:.4f}")
    print(f"    Macro Recall     : {macro_recall:.4f}")
    print(f"    Macro F1         : {macro_f1:.4f}")
    print(f"    Weighted F1      : {weighted_f1:.4f}")
    
    print(f"\n  Entity-Level Metrics (strict span matching):")
    print(f"    Entity Precision : {entity_metrics['precision']:.4f}")
    print(f"    Entity Recall    : {entity_metrics['recall']:.4f}")
    print(f"    Entity F1        : {entity_metrics['f1']:.4f}")
    print(f"    Correct/Total    : {entity_metrics['correct_entities']}/{entity_metrics['total_true_entities']}")
    
    print(f"\n  Per-Entity-Type Metrics:")
    print(f"    {'Entity Type':<20s} {'Prec':>6s} {'Recall':>7s} {'F1':>6s} {'Support':>8s}")
    print(f"    {'-'*50}")
    for etype, m in sorted(entity_metrics['per_entity_type'].items()):
        print(f"    {etype:<20s} {m['precision']:>6.4f} {m['recall']:>7.4f} {m['f1']:>6.4f} {m['support']:>8d}")
    
    print(f"\n  Per-Tag Token-Level Metrics:")
    print(f"    {'Tag':<25s} {'Prec':>6s} {'Recall':>7s} {'F1':>6s} {'Support':>8s}")
    print(f"    {'-'*55}")
    for tag in sorted(per_tag.keys()):
        m = per_tag[tag]
        print(f"    {tag:<25s} {m['precision']:>6.4f} {m['recall']:>7.4f} {m['f1']:>6.4f} {m['support']:>8d}")
    
    # Full sklearn classification report
    print(f"\n  sklearn classification_report (entity tags only):")
    print(classification_report(true_tags, pred_tags, labels=entity_tag_names, zero_division=0))
    
    # Save
    results_path = os.path.join(save_dir, 'evaluation_results.json')
    with open(results_path, 'w') as f:
        json.dump(results, f, indent=2)
    print(f"  Evaluation results saved to: {results_path}")
    
    return results


# ============================================================
#  Training
# ============================================================

def train_model(model, train_loader, val_loader, device, epochs, lr,
                id2tag, tag2id, save_dir):
    """Training loop with validation."""
    
    criterion = nn.CrossEntropyLoss(ignore_index=-100)
    optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=0.01)
    
    total_steps = len(train_loader) * epochs
    scheduler = get_linear_schedule_with_warmup(
        optimizer,
        num_warmup_steps=int(total_steps * 0.1),
        num_training_steps=total_steps
    )
    
    best_f1 = 0.0
    entity_tag_names = sorted(set(tag for tag in id2tag.values() if tag != 'O'))
    training_history = []
    
    for epoch in range(epochs):
        # --- Train ---
        model.train()
        total_loss = 0
        
        for batch_idx, (input_ids, attention_mask, labels) in enumerate(train_loader):
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            labels = labels.to(device)
            
            optimizer.zero_grad()
            logits = model(input_ids, attention_mask)
            
            # Reshape: (batch*seq_len, num_tags) vs (batch*seq_len)
            loss = criterion(logits.view(-1, logits.size(-1)), labels.view(-1))
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            scheduler.step()
            
            total_loss += loss.item()
            
            if (batch_idx + 1) % 20 == 0:
                print(f"    Batch {batch_idx+1}/{len(train_loader)} | Loss: {loss.item():.4f}")
        
        avg_loss = total_loss / len(train_loader)
        
        # --- Quick validation ---
        true_tags, pred_tags = collect_predictions(model, val_loader, device, id2tag)
        
        token_acc = accuracy_score(true_tags, pred_tags)
        micro_f1 = f1_score(true_tags, pred_tags, labels=entity_tag_names,
                            average='micro', zero_division=0)
        macro_f1 = f1_score(true_tags, pred_tags, labels=entity_tag_names,
                            average='macro', zero_division=0)
        
        epoch_metrics = {
            'epoch': epoch + 1,
            'train_loss': round(avg_loss, 4),
            'token_accuracy': round(token_acc, 4),
            'micro_f1': round(micro_f1, 4),
            'macro_f1': round(macro_f1, 4),
        }
        training_history.append(epoch_metrics)
        
        print(f"\n  Epoch {epoch+1}/{epochs} | Loss: {avg_loss:.4f} | Token Acc: {token_acc:.4f} | Micro F1: {micro_f1:.4f} | Macro F1: {macro_f1:.4f}")
        
        # Save best model
        if micro_f1 > best_f1:
            best_f1 = micro_f1
            torch.save({
                'model_state_dict': model.state_dict(),
                'epoch': epoch,
                'best_f1': best_f1,
            }, os.path.join(save_dir, 'model.pt'))
            print(f"  ★ New best model saved (Micro F1={best_f1:.4f})")
        print()
    
    # Save training history
    with open(os.path.join(save_dir, 'training_history.json'), 'w') as f:
        json.dump(training_history, f, indent=2)
    
    return best_f1


# ============================================================
#  Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Train Slot Extractor (NER)')
    parser.add_argument('--data', type=str,
                        default='shopping_queries_dataset/slotannotationdataset.xlsx')
    parser.add_argument('--bert_model', type=str,
                        default='bert-base-multilingual-uncased')
    parser.add_argument('--max_length', type=int, default=64)
    parser.add_argument('--batch_size', type=int, default=16)
    parser.add_argument('--epochs', type=int, default=8)
    parser.add_argument('--lr', type=float, default=3e-5)
    parser.add_argument('--test_size', type=float, default=0.15)
    parser.add_argument('--save_dir', type=str, default='models/slot_extractor')
    args = parser.parse_args()
    
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Device: {device}")
    print()
    
    # 1. Load data
    print("[1/5] Loading dataset...")
    tokens_list, labels_list, all_tags = load_slot_dataset(args.data)
    
    # Build tag maps
    tag2id = {tag: i for i, tag in enumerate(all_tags)}
    id2tag = {i: tag for tag, i in tag2id.items()}
    print(f"\n  Tag map ({len(tag2id)} tags):")
    for tag, idx in tag2id.items():
        print(f"    {idx:2d} → {tag}")
    
    # 2. Split
    print("\n[2/5] Splitting train/val...")
    train_tokens, val_tokens, train_labels, val_labels = train_test_split(
        tokens_list, labels_list, test_size=args.test_size, random_state=42
    )
    print(f"  Train: {len(train_tokens)} | Val: {len(val_tokens)}")
    
    # 3. Tokenize and create datasets
    print("\n[3/5] Creating datasets...")
    tokenizer = BertTokenizerFast.from_pretrained(args.bert_model)
    
    train_dataset = SlotDataset(train_tokens, train_labels, tokenizer, tag2id, args.max_length)
    val_dataset = SlotDataset(val_tokens, val_labels, tokenizer, tag2id, args.max_length)
    
    train_loader = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=args.batch_size)
    
    print(f"  Train batches: {len(train_loader)} | Val batches: {len(val_loader)}")
    
    # 4. Train
    print(f"\n[4/5] Training ({args.epochs} epochs, lr={args.lr}, batch_size={args.batch_size})...")
    os.makedirs(args.save_dir, exist_ok=True)
    
    model = SlotExtractor(
        bert_model_name=args.bert_model,
        num_tags=len(tag2id)
    ).to(device)
    
    print(f"  Model parameters: {sum(p.numel() for p in model.parameters()):,}")
    print()
    
    start = time.time()
    best_f1 = train_model(model, train_loader, val_loader, device,
                           args.epochs, args.lr, id2tag, tag2id, args.save_dir)
    elapsed = time.time() - start
    
    # 5. Final Evaluation — load best model and run comprehensive metrics
    print("\n[5/5] Running comprehensive evaluation on best model...")
    checkpoint = torch.load(os.path.join(args.save_dir, 'model.pt'), map_location=device)
    model.load_state_dict(checkpoint['model_state_dict'])
    
    eval_results = evaluate_slot_model(model, val_loader, device, id2tag, tag2id, args.save_dir)
    
    # Save tag map
    with open(os.path.join(args.save_dir, 'tag_map.json'), 'w') as f:
        json.dump(tag2id, f, indent=2)
    
    # Save config
    config = {
        'bert_model': args.bert_model,
        'max_length': args.max_length,
        'num_tags': len(tag2id),
        'tag_names': all_tags,
        'best_micro_f1': best_f1,
        'training_time_seconds': round(elapsed, 1),
        'dataset': args.data,
        'dataset_size': len(tokens_list),
        'train_size': len(train_tokens),
        'val_size': len(val_tokens),
        'final_token_metrics': eval_results['token_level'],
        'final_entity_metrics': {
            'precision': eval_results['entity_level']['precision'],
            'recall': eval_results['entity_level']['recall'],
            'f1': eval_results['entity_level']['f1'],
        },
    }
    with open(os.path.join(args.save_dir, 'config.json'), 'w') as f:
        json.dump(config, f, indent=2)
    
    print(f"\n{'='*70}")
    print(f"  ✓ TRAINING COMPLETE")
    print(f"{'='*70}")
    print(f"  Time              : {elapsed/60:.1f} minutes")
    print(f"  Best Micro F1     : {best_f1:.4f}")
    print(f"  Token Accuracy    : {eval_results['token_level']['accuracy']:.4f}")
    print(f"  Entity-Level F1   : {eval_results['entity_level']['f1']:.4f}")
    print(f"  Saved to          : {args.save_dir}/")
    print(f"    model.pt, tag_map.json, config.json,")
    print(f"    evaluation_results.json, training_history.json")


if __name__ == '__main__':
    main()
