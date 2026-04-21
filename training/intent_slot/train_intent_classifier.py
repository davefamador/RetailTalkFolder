"""
Train Intent Classifier
Fine-tunes bert-base-multilingual-uncased for multi-label intent classification.

Dataset: shopping_queries_dataset/IntentDataset_cleaned.xlsx
Labels:  single_search, multi_search, filtered_search, free_form (multi-label, 0/1)
Output:  models/intent_classifier/model.pt + label_map.json + evaluation_results.json

Usage:
  python train_intent_classifier.py
  python train_intent_classifier.py --epochs 10 --batch_size 32 --lr 2e-5
"""

import os
import json
import torch
import torch.nn as nn
import numpy as np
import openpyxl
import matplotlib.pyplot as plt
import seaborn as sns
from torch.utils.data import DataLoader, TensorDataset
from transformers import BertModel, BertTokenizer, get_linear_schedule_with_warmup
from sklearn.model_selection import train_test_split
from sklearn.metrics import (
    classification_report, f1_score, accuracy_score,
    precision_score, recall_score, hamming_loss,
    confusion_matrix, multilabel_confusion_matrix
)
from nltk.translate.bleu_score import corpus_bleu, SmoothingFunction
import argparse
import time


# ============================================================
#  Model Architecture
# ============================================================

class IntentClassifier(nn.Module):
    """
    BERT + classification head for multi-label intent classification.
    
    Architecture:
      Input text → BERT tokenizer → BERT encoder (768-dim per token)
      → [CLS] token (768-dim) → Dropout → Linear(768,256) → ReLU
      → Dropout → Linear(256, num_intents) → Sigmoid
    
    Sigmoid (not Softmax) because a query can have MULTIPLE intents:
      e.g., "shoes and bag under 3000" → multi_search=1, filtered_search=1
    """
    
    def __init__(self, bert_model_name='bert-base-multilingual-uncased', num_intents=4):
        super().__init__()
        self.bert = BertModel.from_pretrained(bert_model_name) # Model Architecture until fc2
        self.dropout = nn.Dropout(0.3)
        self.fc1 = nn.Linear(768, 256)
        self.relu = nn.ReLU()
        self.fc2 = nn.Linear(256, num_intents)
    
    def forward(self, input_ids, attention_mask):
        outputs = self.bert(input_ids=input_ids, attention_mask=attention_mask)
        cls_output = outputs.last_hidden_state[:, 0, :]  # [CLS] token
        x = self.dropout(cls_output)
        x = self.relu(self.fc1(x))
        x = self.dropout(x)
        logits = self.fc2(x)
        return logits  # raw logits, apply sigmoid in loss/inference


# ============================================================
#  Data Loading
# ============================================================

def load_intent_dataset(filepath):
    """Load intent dataset from xlsx. Returns queries and label arrays."""
    print(f"  Loading {filepath}...")
    wb = openpyxl.load_workbook(filepath)
    ws = wb.active
    
    queries = []
    labels = []
    label_names = ['single_search', 'multi_search', 'filtered_search', 'free_form']
    
    for r in range(2, ws.max_row + 1):
        query = ws.cell(r, 1).value
        if not query:
            continue
        query = str(query).strip()
        row_labels = []
        for c in range(2, 6):  # columns 2-5
            val = ws.cell(r, c).value
            row_labels.append(int(val) if val else 0)
        queries.append(query)
        labels.append(row_labels)
    
    print(f"  Loaded {len(queries)} queries")
    print(f"  Label distribution:")
    labels_arr = np.array(labels)
    for i, name in enumerate(label_names):
        print(f"    {name}: {labels_arr[:, i].sum()}")
    
    return queries, np.array(labels, dtype=np.float32), label_names


def tokenize_queries(queries, tokenizer, max_length=128):
    """Tokenize all queries with BERT tokenizer."""
    print(f"  Tokenizing {len(queries)} queries (max_length={max_length})...")
    encodings = tokenizer(
        queries,
        padding='max_length',
        truncation=True,
        max_length=max_length,
        return_tensors='pt'
    )
    return encodings['input_ids'], encodings['attention_mask']


# ============================================================
#  Evaluation
# ============================================================

def evaluate_intent_model(model, val_loader, device, label_names, save_dir):
    """
    Comprehensive evaluation of intent classifier.
    
    Metrics computed:
    - Accuracy (exact match ratio — all labels must match)
    - Hamming Loss (fraction of wrong labels)
    - Per-label: Precision, Recall, F1
    - Micro/Macro averages: Precision, Recall, F1
    - Per-label Confusion Matrix (TP, FP, FN, TN)
    """
    model.eval()
    all_preds = []
    all_labels = []
    all_probs = []
    
    with torch.no_grad():
        for input_ids, attention_mask, labels in val_loader:
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            
            logits = model(input_ids, attention_mask)
            probs = torch.sigmoid(logits).cpu().numpy()
            preds = (probs > 0.5).astype(int)
            
            all_probs.append(probs)
            all_preds.append(preds)
            all_labels.append(labels.numpy())
    
    all_preds = np.vstack(all_preds)
    all_labels = np.vstack(all_labels).astype(int)
    all_probs = np.vstack(all_probs)
    
    # ---- Exact Match Accuracy ----
    # ALL labels must match for a query to be "correct"
    exact_match = accuracy_score(all_labels, all_preds)
    
    # ---- Hamming Loss ----
    # Fraction of individual labels that are wrong
    h_loss = hamming_loss(all_labels, all_preds)
    
    # ---- Per-Label Metrics ----
    per_label = {}
    for i, name in enumerate(label_names):
        precision = precision_score(all_labels[:, i], all_preds[:, i], zero_division=0)
        recall = recall_score(all_labels[:, i], all_preds[:, i], zero_division=0)
        f1 = f1_score(all_labels[:, i], all_preds[:, i], zero_division=0)
        acc = accuracy_score(all_labels[:, i], all_preds[:, i])
        
        # Confusion matrix for this label
        cm = confusion_matrix(all_labels[:, i], all_preds[:, i], labels=[0, 1])
        tn, fp, fn, tp = cm.ravel()
        
        per_label[name] = {
            'precision': round(precision, 4),
            'recall': round(recall, 4),
            'f1': round(f1, 4),
            'accuracy': round(acc, 4),
            'support': int(all_labels[:, i].sum()),
            'confusion_matrix': {
                'TP': int(tp), 'FP': int(fp),
                'FN': int(fn), 'TN': int(tn)
            }
        }
    
    # ---- Micro / Macro Averages ----
    micro_precision = precision_score(all_labels, all_preds, average='micro', zero_division=0)
    micro_recall = recall_score(all_labels, all_preds, average='micro', zero_division=0)
    micro_f1 = f1_score(all_labels, all_preds, average='micro', zero_division=0)
    
    macro_precision = precision_score(all_labels, all_preds, average='macro', zero_division=0)
    macro_recall = recall_score(all_labels, all_preds, average='macro', zero_division=0)
    macro_f1 = f1_score(all_labels, all_preds, average='macro', zero_division=0)
    
    weighted_f1 = f1_score(all_labels, all_preds, average='weighted', zero_division=0)
    samples_f1 = f1_score(all_labels, all_preds, average='samples', zero_division=0)
    
    # ---- BLEU Score ----
    # Converts multi-label vectors to label name sequences and computes BLEU
    # e.g., true=[1,1,0,0] → ["single_search", "multi_search"]
    #        pred=[1,0,0,0] → ["single_search"]
    smoother = SmoothingFunction().method1
    references = []
    hypotheses = []
    for i in range(len(all_labels)):
        ref_labels = [label_names[j] for j in range(len(label_names)) if all_labels[i][j] == 1]
        pred_labels = [label_names[j] for j in range(len(label_names)) if all_preds[i][j] == 1]
        if not ref_labels:
            ref_labels = ['none']
        if not pred_labels:
            pred_labels = ['none']
        references.append([ref_labels])  # corpus_bleu expects list of references
        hypotheses.append(pred_labels)
    
    bleu_score = corpus_bleu(references, hypotheses, smoothing_function=smoother)
    
    # ---- Build Results ----
    results = {
        'overall': {
            'exact_match_accuracy': round(exact_match, 4),
            'hamming_loss': round(h_loss, 4),
            'bleu_score': round(bleu_score, 4),
            'micro_precision': round(micro_precision, 4),
            'micro_recall': round(micro_recall, 4),
            'micro_f1': round(micro_f1, 4),
            'macro_precision': round(macro_precision, 4),
            'macro_recall': round(macro_recall, 4),
            'macro_f1': round(macro_f1, 4),
            'weighted_f1': round(weighted_f1, 4),
            'samples_f1': round(samples_f1, 4),
        },
        'per_label': per_label,
        'total_samples': len(all_labels),
    }
    
    # ---- Print Report ----
    print("\n" + "=" * 70)
    print("  EVALUATION RESULTS — INTENT CLASSIFIER")
    print("=" * 70)
    
    print(f"\n  Overall Metrics:")
    print(f"    Exact Match Accuracy : {exact_match:.4f}")
    print(f"    Hamming Loss         : {h_loss:.4f}")
    print(f"    BLEU Score           : {bleu_score:.4f}")
    print(f"    Micro F1             : {micro_f1:.4f}")
    print(f"    Macro F1             : {macro_f1:.4f}")
    print(f"    Weighted F1          : {weighted_f1:.4f}")
    print(f"    Samples F1           : {samples_f1:.4f}")
    
    print(f"\n  Per-Label Metrics:")
    print(f"    {'Label':<20s} {'Prec':>6s} {'Recall':>7s} {'F1':>6s} {'Acc':>6s} {'Support':>8s}")
    print(f"    {'-'*55}")
    for name in label_names:
        m = per_label[name]
        print(f"    {name:<20s} {m['precision']:>6.4f} {m['recall']:>7.4f} {m['f1']:>6.4f} {m['accuracy']:>6.4f} {m['support']:>8d}")
    
    print(f"\n  Confusion Matrices (per label):")
    for name in label_names:
        cm = per_label[name]['confusion_matrix']
        print(f"\n    {name}:")
        print(f"                  Predicted 0    Predicted 1")
        print(f"      Actual 0    TN={cm['TN']:<10d} FP={cm['FP']}")
        print(f"      Actual 1    FN={cm['FN']:<10d} TP={cm['TP']}")
    
    # ---- Full sklearn report ----
    print(f"\n  sklearn classification_report:")
    print(classification_report(
        all_labels, all_preds,
        target_names=label_names,
        zero_division=0
    ))
    
    # ---- Confusion Matrix Plots ----
    # For multi-label: one binary confusion matrix per label (2x2: 0/1 vs 0/1)
    mcm = multilabel_confusion_matrix(all_labels, all_preds)  # (num_labels, 2, 2)
    fig, axes = plt.subplots(1, len(label_names), figsize=(4 * len(label_names), 4))
    if len(label_names) == 1:
        axes = [axes]
    for i, (name, cm_i) in enumerate(zip(label_names, mcm)):
        sns.heatmap(
            cm_i, annot=True, fmt='d', cmap='Blues',
            xticklabels=['Pred 0', 'Pred 1'],
            yticklabels=['True 0', 'True 1'],
            ax=axes[i]
        )
        axes[i].set_title(name, fontsize=10)
    fig.suptitle('Intent Classifier — Per-Label Confusion Matrices', fontsize=12, y=1.02)
    plt.tight_layout()
    cm_path = os.path.join(save_dir, 'confusion_matrix.png')
    plt.savefig(cm_path, dpi=150, bbox_inches='tight')
    plt.close()
    print(f"  Confusion matrix plot saved to: {cm_path}")

    # ---- Save to JSON ----
    results_path = os.path.join(save_dir, 'evaluation_results.json')
    with open(results_path, 'w') as f:
        json.dump(results, f, indent=2)
    print(f"  Evaluation results saved to: {results_path}")

    return results


# ============================================================
#  Training
# ============================================================

def train_model(model, train_loader, val_loader, device, epochs, lr, label_names, save_dir,
                start_epoch=0, optimizer=None, scheduler=None, best_f1=0.0, training_history=None):
    """Training loop with validation and checkpoint support."""
    
    # BCEWithLogitsLoss for multi-label (combines sigmoid + binary cross-entropy)
    criterion = nn.BCEWithLogitsLoss()
    
    if optimizer is None:
        optimizer = torch.optim.AdamW(model.parameters(), lr=lr, weight_decay=0.01)
    
    total_steps = len(train_loader) * epochs
    if scheduler is None:
        scheduler = get_linear_schedule_with_warmup(
            optimizer,
            num_warmup_steps=int(total_steps * 0.1),
            num_training_steps=total_steps
        )
    
    if training_history is None:
        training_history = []
    
    for epoch in range(start_epoch, epochs):
        # --- Train ---
        model.train()
        total_loss = 0
        
        for batch_idx, (input_ids, attention_mask, labels) in enumerate(train_loader):
            input_ids = input_ids.to(device)
            attention_mask = attention_mask.to(device)
            labels = labels.to(device)
            
            optimizer.zero_grad()
            logits = model(input_ids, attention_mask)
            loss = criterion(logits, labels)
            loss.backward()
            torch.nn.utils.clip_grad_norm_(model.parameters(), max_norm=1.0)
            optimizer.step()
            scheduler.step()
            
            total_loss += loss.item()
            
            if (batch_idx + 1) % 20 == 0:
                print(f"    Batch {batch_idx+1}/{len(train_loader)} | Loss: {loss.item():.4f}")
        
        avg_loss = total_loss / len(train_loader)
        
        # --- Quick validation for epoch summary ---
        model.eval()
        all_preds = []
        all_labels = []
        
        with torch.no_grad():
            for input_ids, attention_mask, labels in val_loader:
                input_ids = input_ids.to(device)
                attention_mask = attention_mask.to(device)
                
                logits = model(input_ids, attention_mask)
                probs = torch.sigmoid(logits)
                preds = (probs > 0.5).int().cpu().numpy()
                
                all_preds.append(preds)
                all_labels.append(labels.numpy())
        
        all_preds = np.vstack(all_preds)
        all_labels = np.vstack(all_labels).astype(int)
        
        # Per-label F1
        f1_per_label = []
        for i, name in enumerate(label_names):
            f1 = f1_score(all_labels[:, i], all_preds[:, i], zero_division=0)
            f1_per_label.append(f1)
        
        macro_f1 = np.mean(f1_per_label)
        exact_match = accuracy_score(all_labels, all_preds)
        h_loss = hamming_loss(all_labels, all_preds)
        
        epoch_metrics = {
            'epoch': epoch + 1,
            'train_loss': round(avg_loss, 4),
            'macro_f1': round(macro_f1, 4),
            'exact_match': round(exact_match, 4),
            'hamming_loss': round(h_loss, 4),
            'per_label_f1': {name: round(f1_per_label[i], 4) for i, name in enumerate(label_names)}
        }
        training_history.append(epoch_metrics)
        
        print(f"\n  Epoch {epoch+1}/{epochs} | Loss: {avg_loss:.4f} | Macro F1: {macro_f1:.4f} | Exact Match: {exact_match:.4f} | Hamming: {h_loss:.4f}")
        for i, name in enumerate(label_names):
            print(f"    {name:20s}: F1={f1_per_label[i]:.4f}")
        
        # Save best model
        if macro_f1 > best_f1:
            best_f1 = macro_f1
            torch.save({
                'model_state_dict': model.state_dict(),
                'epoch': epoch,
                'best_f1': best_f1,
            }, os.path.join(save_dir, 'model.pt'))
            print(f"  ★ New best model saved (F1={best_f1:.4f})")
        
        # Save training checkpoint (for resuming)
        torch.save({
            'epoch': epoch + 1,  # next epoch to run
            'model_state_dict': model.state_dict(),
            'optimizer_state_dict': optimizer.state_dict(),
            'scheduler_state_dict': scheduler.state_dict(),
            'best_f1': best_f1,
            'training_history': training_history,
        }, os.path.join(save_dir, 'checkpoint.pt'))
        print(f"  💾 Checkpoint saved (epoch {epoch+1}/{epochs})")
        print()
    
    # Save training history
    with open(os.path.join(save_dir, 'training_history.json'), 'w') as f:
        json.dump(training_history, f, indent=2)
    
    return best_f1


# ============================================================
#  Main
# ============================================================

def main():
    parser = argparse.ArgumentParser(description='Train Intent Classifier')
    parser.add_argument('--data', type=str,
                        default='shopping_queries_dataset/IntentDataset_cleaned.xlsx')
    parser.add_argument('--bert_model', type=str,
                        default='bert-base-multilingual-uncased')
    parser.add_argument('--max_length', type=int, default=128)
    parser.add_argument('--batch_size', type=int, default=16)
    parser.add_argument('--epochs', type=int, default=5)
    parser.add_argument('--lr', type=float, default=2e-5)
    parser.add_argument('--test_size', type=float, default=0.15)
    parser.add_argument('--save_dir', type=str, default='models/intent_classifier')
    parser.add_argument('--resume', action='store_true',
                        help='Resume training from the last saved checkpoint')
    args = parser.parse_args()
    
    device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
    print(f"Device: {device}")
    print()
    
    # 1. Load data
    print("[1/5] Loading dataset...")
    queries, labels, label_names = load_intent_dataset(args.data)
    
    # 2. Split
    print("\n[2/5] Splitting train/val...")
    train_q, val_q, train_l, val_l = train_test_split(
        queries, labels, test_size=args.test_size, random_state=42
    )
    print(f"  Train: {len(train_q)} | Val: {len(val_q)}")
    
    # 3. Tokenize
    print("\n[3/5] Tokenizing...")
    tokenizer = BertTokenizer.from_pretrained(args.bert_model)
    
    train_ids, train_masks = tokenize_queries(train_q, tokenizer, args.max_length)
    val_ids, val_masks = tokenize_queries(val_q, tokenizer, args.max_length)
    
    train_labels = torch.tensor(train_l, dtype=torch.float32)
    val_labels = torch.tensor(val_l, dtype=torch.float32)
    
    train_dataset = TensorDataset(train_ids, train_masks, train_labels)
    val_dataset = TensorDataset(val_ids, val_masks, val_labels)
    
    train_loader = DataLoader(train_dataset, batch_size=args.batch_size, shuffle=True)
    val_loader = DataLoader(val_dataset, batch_size=args.batch_size)
    
    # 4. Train
    print(f"\n[4/5] Training ({args.epochs} epochs, lr={args.lr}, batch_size={args.batch_size})...")
    os.makedirs(args.save_dir, exist_ok=True)
    
    model = IntentClassifier(
        bert_model_name=args.bert_model,
        num_intents=len(label_names)
    ).to(device)
    
    print(f"  Model parameters: {sum(p.numel() for p in model.parameters()):,}")
    print()
    
    # Checkpoint resume support
    start_epoch = 0
    resume_best_f1 = 0.0
    resume_optimizer = None
    resume_scheduler = None
    resume_history = None
    
    checkpoint_path = os.path.join(args.save_dir, 'checkpoint.pt')
    if args.resume and os.path.exists(checkpoint_path):
        print(f"  ⏩ Resuming from checkpoint: {checkpoint_path}")
        ckpt = torch.load(checkpoint_path, map_location=device, weights_only=False)
        model.load_state_dict(ckpt['model_state_dict'])
        start_epoch = ckpt['epoch']
        resume_best_f1 = ckpt['best_f1']
        resume_history = ckpt.get('training_history', [])
        
        # Rebuild optimizer & scheduler then load their states
        total_steps = len(train_loader) * args.epochs
        resume_optimizer = torch.optim.AdamW(model.parameters(), lr=args.lr, weight_decay=0.01)
        resume_scheduler = get_linear_schedule_with_warmup(
            resume_optimizer,
            num_warmup_steps=int(total_steps * 0.1),
            num_training_steps=total_steps
        )
        resume_optimizer.load_state_dict(ckpt['optimizer_state_dict'])
        resume_scheduler.load_state_dict(ckpt['scheduler_state_dict'])
        
        print(f"  ⏩ Resuming from epoch {start_epoch + 1}/{args.epochs} (best F1 so far: {resume_best_f1:.4f})")
    elif args.resume:
        print(f"  ⚠ No checkpoint found at {checkpoint_path}, starting from scratch.")
    
    start = time.time()
    best_f1 = train_model(model, train_loader, val_loader, device,
                           args.epochs, args.lr, label_names, args.save_dir,
                           start_epoch=start_epoch, optimizer=resume_optimizer,
                           scheduler=resume_scheduler, best_f1=resume_best_f1,
                           training_history=resume_history)
    elapsed = time.time() - start
    
    # 5. Final Evaluation — load best model and run comprehensive metrics
    print("\n[5/5] Running comprehensive evaluation on best model...")
    checkpoint = torch.load(os.path.join(args.save_dir, 'model.pt'), map_location=device, weights_only=False)
    model.load_state_dict(checkpoint['model_state_dict'])
    
    eval_results = evaluate_intent_model(model, val_loader, device, label_names, args.save_dir)
    
    # Save label map
    label_map = {name: i for i, name in enumerate(label_names)}
    with open(os.path.join(args.save_dir, 'label_map.json'), 'w') as f:
        json.dump(label_map, f, indent=2)
    
    # Save config
    config = {
        'bert_model': args.bert_model,
        'max_length': args.max_length,
        'num_intents': len(label_names),
        'label_names': label_names,
        'best_f1': best_f1,
        'training_time_seconds': round(elapsed, 1),
        'dataset': args.data,
        'dataset_size': len(queries),
        'train_size': len(train_q),
        'val_size': len(val_q),
        'final_metrics': eval_results['overall'],
    }
    with open(os.path.join(args.save_dir, 'config.json'), 'w') as f:
        json.dump(config, f, indent=2)
    
    print(f"\n{'='*70}")
    print(f"  ✓ TRAINING COMPLETE")
    print(f"{'='*70}")
    print(f"  Time        : {elapsed/60:.1f} minutes")
    print(f"  Best Macro F1: {best_f1:.4f}")
    print(f"  Exact Match  : {eval_results['overall']['exact_match_accuracy']:.4f}")
    print(f"  Hamming Loss : {eval_results['overall']['hamming_loss']:.4f}")
    print(f"  Saved to     : {args.save_dir}/")
    print(f"    model.pt, label_map.json, config.json,")
    print(f"    evaluation_results.json, training_history.json")


if __name__ == '__main__':
    main()
